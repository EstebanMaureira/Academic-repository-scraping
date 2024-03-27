 
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Nov 25 20:58:36 2023

@author: estebanmaureiravenegas
"""

# Configuracion inicial de Selenium y WebDriver
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import requests
from bs4 import BeautifulSoup
import csv
import unicodedata
from openpyxl.utils import get_column_letter
import time
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException
from openpyxl import Workbook, load_workbook
import random
import re
 


# =============================================================================
# # Funciones auxiliares
# =============================================================================


# =============================================================================
# pausas
# =============================================================================
 
cache_paginas = {}

def obtener_html(url, driver):
    if url in cache_paginas:
        return cache_paginas[url]
    else:
        driver.get(url)
        esperar_hasta_cargar_completamente(driver)
        time.sleep(2)  # Espera dos segundos antes de continuar
        html = driver.page_source
        cache_paginas[url] = html
        return html




def normalizar_nombre_archivo(nombre):
    return "".join(c for c in nombre if c.isalnum() or c in (' ', '-', '_')).rstrip().replace(' ', '_')

def normalizar_nombre_autor(nombre):
    # Normalización Unicode para manejar caracteres no ASCII
    nombre_normalizado = unicodedata.normalize('NFKD', nombre).encode('ascii', 'ignore').decode('ascii')
    
    # Conservando caracteres alfanuméricos, espacios, apóstrofes y guiones (comunes en nombres)
    nombre_limpio = "".join(c for c in nombre_normalizado if c.isalnum() or c in (' ', '-', "'"))
    
    # Eliminando espacios extra al principio y al final
    return nombre_limpio.strip()

# =============================================================================
# 
# =============================================================================
from urllib.parse import urlparse, unquote


def obtener_extension(url):
    """
    Extrae la extensión del archivo de un URL.
    """
    # Decodificar el URL para manejar espacios y caracteres especiales
    ruta = urlparse(unquote(url)).path
    return os.path.splitext(ruta)[1]
# =============================================================================
# 
# =============================================================================

def es_video(url):
    # Una forma simple de determinar si es un video podría ser por la extensión en la URL
    extensiones_video = ['.mp4', '.avi', '.mov', '.wmv', '.flv']
    return any(url.endswith(ext) for ext in extensiones_video)


def descargar_archivo(url, nombre_archivo, directorio_destino):
    if not os.path.exists(directorio_destino):
        os.makedirs(directorio_destino)   
    ruta_completa = os.path.join(directorio_destino, nombre_archivo)
    
    # Obtener el tamaño del archivo del encabezado si está disponible
    respuesta = requests.head(url, allow_redirects=True, verify=False)
    size_header = respuesta.headers.get('content-length')
    if size_header:
        size_total_mb = int(size_header) / (1024 * 1024)  # Convertir a MB
    else:
        size_total_mb = 'Desconocido'

    # Descargar el archivo
    respuesta = requests.get(url, stream=True, verify=False)
    
    if respuesta.status_code == 200:
        size_descargado = 0
        with open(ruta_completa, 'wb') as f:
            for chunk in respuesta.iter_content(1024):
                f.write(chunk)
                size_descargado += len(chunk)
        size_descargado_mb = size_descargado / (1024 * 1024)  # Convertir a MB
        
        # Pausa aleatoria entre 1 y 5 segundos para simular un comportamiento humano más natural
        time.sleep(random.uniform(1, 5))
        
        # Formatear el mensaje de estado
        if size_total_mb != 'Desconocido':
            status_doc = f"{size_descargado_mb:.2f} MB descargados de {size_total_mb:.2f} MB"
        else:
            status_doc = f"{size_descargado_mb:.2f} MB descargados de tamaño desconocido"
        
        return ruta_completa, status_doc  # Retorna la ruta completa y el estado
    else:
        print(f"Error al descargar el archivo: {respuesta.status_code}")
        return None, f"Error al descargar el archivo: {respuesta.status_code}"

# =============================================================================
# 
# =============================================================================

def esperar_descarga_completa(ruta_archivo):
    while not os.path.exists(ruta_archivo):
        time.sleep(1)  # Espera a que el archivo exista

    size_antiguo = -1
    while True:
        size_actual = os.path.getsize(ruta_archivo)
        if size_actual == size_antiguo:
            break  # Salir del bucle si el tamaño del archivo deja de cambiar
        else:
            size_antiguo = size_actual
            time.sleep(1)  # Espera un segundo antes de la próxima comprobación
# =============================================================================
# carga completa
# =============================================================================
def esperar_hasta_cargar_completamente(driver, timeout=30):
    """ Espera hasta que la página esté completamente cargada. """
    inicio = time.time()
    while True:
        # Comprobar el estado de 'document.readyState'
        estado_pagina = driver.execute_script("return document.readyState;")
        if estado_pagina == "complete":
            break
        elif time.time() - inicio > timeout:
            raise TimeoutError("La página no se cargó en el tiempo esperado.")
        time.sleep(0.5)  # Esperar un poco antes de la próxima comprobación
def limpiar_cache(driver):
    driver.execute_script("window.localStorage.clear();")
    driver.execute_script("window.sessionStorage.clear();")

# Llamar a esta función dentro de tu bucle principal o donde sea necesario

class ServerError(Exception):
    """Excepción personalizada para errores relacionados con el servidor o la página web."""
    pass
# =============================================================================
# =============================================================================
# # 
# =============================================================================
# =============================================================================
def obtener_offset_del_enlace(enlace):
    match = re.search(r'offset=(\d+)', enlace)
    return int(match.group(1)) if match else 0
# =============================================================================
# =============================================================================
# # 
# =============================================================================
# =============================================================================

def login_and_scrape(offset_inicial):
    driver = None
    contador_iteraciones = 0
    wb = None  # Inicializa wb aquí
    try:
        # =============================================================================
        # # Iniciar sesion
        # =============================================================================
        firefox_options = Options()
        firefox_options.add_argument('--ignore-ssl-errors=true')
        firefox_options.add_argument('--ignore-certificate-errors')
        firefox_options.accept_insecure_certs = True  # Esta línea hace lo mismo que las anteriores

        service = Service(r'ruta')  # Actualizar con la ruta correcta
        # Configuracion de descarga
        download_path_doc = r"doc"
        download_path_mtd = r"metadato"
        driver = webdriver.Firefox(service=service, options=firefox_options)
        driver.set_page_load_timeout(600)
        # =============================================================================
        # # Acceso a la página de títulos
        # =============================================================================
        url_base = 'url institucional'
        obtener_html(url_base, driver)
        esperar_hasta_cargar_completamente(driver)

# =============================================================================
# =============================================================================
# #         # Cargar el archivo Excel existente o crear uno nuevo       
# =============================================================================
# =============================================================================
  
        ruta_archivo_excel = 'Planilla_Proyecto.xlsx'
        offset_inicial = 0
        indice = 1  # Valor predeterminado para 'indice'
        
        # Revisar si el archivo Excel ya existe
        if os.path.exists(ruta_archivo_excel):
            wb = load_workbook(ruta_archivo_excel)
            ws = wb.active
        
            # Determinar el 'indice' y 'offset_inicial' basados en el contenido del archivo
            if ws.max_row > 1:
                ultimo_indice = ws.cell(row=ws.max_row, column=1).value
                indice = ultimo_indice + 1 if ultimo_indice is not None else 1
        
                enlace_ultimo_registro = ws.cell(row=ws.max_row, column=8).value
                if enlace_ultimo_registro and 'offset=' in enlace_ultimo_registro:
                    offset_inicial = obtener_offset_del_enlace(enlace_ultimo_registro)
        
            wb.close()
        else:
            # Crear un nuevo libro y hoja de Excel si el archivo no existe
            wb = Workbook()
            ws = wb.active
            ws.append(['Indice', "dc.identifier", "Titulo", "Autores", "Abstract", "Coleccion", "dc.date.issued", 'URL_Base', 'Url_Repo', "DOI", "Metadato", 'Href_doc', 'Href_doc_2', 'Href_doc_3', 'Href_doc_4','Href_doc_5','Href_doc_6','Href_doc_7','Href_doc_8','Href_doc_9','Href_doc_10','Href_doc_11','Href_doc_12','Href_doc_13','Href_doc_14','Href_doc_15','Href_doc_16','Href_doc_17','Href_doc_18','Href_doc_19','Href_doc_20','Href_doc_21','Href_doc_22','Href_doc_23','Href_doc_24','Href_doc_25','Href_doc_26','Href_doc_27','Href_doc_28','Href_doc_29','Href_doc_30', 'ext_doc', 'ext_doc_2', 'ext_doc_3', 'ext_doc_4','ext_doc_5','ext_doc_6','ext_doc_7','ext_doc_8','ext_doc_9','ext_doc_10','ext_doc_11', 'ext_doc_12', 'ext_doc_13', 'ext_doc_14','ext_doc_15','ext_doc_16','ext_doc_17','ext_doc_18','ext_doc_19','ext_doc_20','ext_doc_21', 'ext_doc_22', 'ext_doc_23', 'ext_doc_24','ext_doc_25','ext_doc_26','ext_doc_27','ext_doc_28','ext_doc_29','ext_doc_30', 'doc_urls', 'doc_urls_2', 'doc_urls_3', 'doc_urls_4','doc_urls_5','doc_urls_6','doc_urls_7','doc_urls_8','doc_urls_9','doc_urls_10','doc_urls_11', 'doc_urls_12', 'doc_urls_13', 'doc_urls_14','doc_urls_15','doc_urls_16','doc_urls_17','doc_urls_18','doc_urls_19','doc_urls_20','doc_urls_21', 'doc_urls_22', 'doc_urls_23', 'doc_urls_24','doc_urls_25','doc_urls_26','doc_urls_27','doc_urls_28','doc_urls_29','doc_urls_30', 'stat_doc', 'stat_doc_2', 'stat_doc_3', 'stat_doc_4','stat_doc_5','stat_doc_6','stat_doc_7','stat_doc_8','stat_doc_9','stat_doc_10','stat_doc_11', 'stat_doc_12', 'stat_doc_13', 'stat_doc_14','stat_doc_15','stat_doc_16','stat_doc_17','stat_doc_18','stat_doc_19','stat_doc_20','stat_doc_21', 'stat_doc_22', 'stat_doc_23', 'stat_doc_24','stat_doc_25','stat_doc_26','stat_doc_27','stat_doc_28','stat_doc_29','stat_doc_30'])

        
        offset = offset_inicial  # Establecer el offset basado en el archivo Excel


# =============================================================================
# =============================================================================
# # motor de la función
# =============================================================================
# =============================================================================

        while offset <= 5879:
            url_base = f'https://___offset={offset}___'
            driver.get(url_base)
            # obtener_html(url_base, driver)
            
   
            esperar_hasta_cargar_completamente(driver)
            elementos_titulos = WebDriverWait(driver, 1).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".col-xs-12.col-sm-12.col-md-9.main-content .artifact-title a")))
            urls_titulos = [elemento.get_attribute('href') for elemento in elementos_titulos]
            for url_titulo in urls_titulos:
                obtener_html(url_titulo, driver)
                # driver.get(url_titulo)
                time.sleep(1) 
                              
                
                # =============================================================================
                # Extraer datos del ti­tulo actual
                # ============================================================================= 
                titulo = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CSS_SELECTOR, "h2.page-header.first-page-header"))).text
                # Verificar si el ti­tulo ya esti¡ registrado en Excel
                
                titulos_registrados = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
                if titulo in titulos_registrados:
                    continue 
    # =============================================================================
    #           Url repo
    # =============================================================================
                url_actual = driver.current_url
    # =============================================================================
    #             abstract
    # =============================================================================
                try:
                    abstract = WebDriverWait(driver, 1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.simple-item-view-description.item-page-field-wrapper.table"))
                    ).text
                    if not abstract.strip():  # Tambii©n verifica si el texto del abstract esti¡ vaci­o
                        abstract = "No disponible"  # O lo que consideres adecuado como marcador de posicion
                except TimeoutException:
                    abstract = "No disponible" 

    # =============================================================================
    #               doi
    # =============================================================================
                try:
                    doi_element = WebDriverWait(driver, 1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.simple-item-view-uri.item-page-field-wrapper.table a"))
                    )
                    doi = doi_element.get_attribute('href')
                    if not doi:  # Verifica si el atributo 'href' esti¡ presente y no esti¡ vaci­o
                        doi = "No disponible"  # O lo que consideres adecuado como marcador de posicion
                except TimeoutException:
                    doi = "No disponible"  # O lo que consideres adecuado como marcador de posicion            
                # =============================================================================
                # # Coleccion
                # =============================================================================
                col_elements = driver.find_elements(By.CSS_SELECTOR, "ul.ds-referenceSet-list li a")
                # Comprobar si se encontraron elementos
                if col_elements:
                    col_links_text = [element.text for element in col_elements if element.text]  # Asegurate de que el texto no esti© vaci­o
                    coleccion = ', '.join(col_links_text)
                else:
                    coleccion = "No disponible"
                
  
                
                # =============================================================================
                # # Autores
                # =============================================================================
                try:
                    elemento_autores = WebDriverWait(driver, 1).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.simple-item-view-authors.item-page-field-wrapper.table"))
                    )
                    autores_divs = elemento_autores.find_elements(By.TAG_NAME, "div")
                    autores = [autor.text for autor in autores_divs if autor.tag_name != "H5" and autor.text.strip() != ""]
                    autores_normalizados = [normalizar_nombre_autor(autor) for autor in autores]
                    autor = "- ".join(autores_normalizados)
                except TimeoutException:
                    autor = "No disponible"
                
                # =============================================================================
                # # Descargar metadatos
                # =============================================================================                     
      
                try:
                    # Intentar encontrar el enlace de metadatos
                    try:
                        metadato_url = driver.find_element(By.CSS_SELECTOR, "div.simple-item-view-show-full.item-page-field-wrapper.table a").get_attribute('href')
                        driver.get(metadato_url)
                    except NoSuchElementException:
                        metadato_url = None
                
                    # Inicializa dc_identifier y metada_ruta
                    dc_identifier = None
                    metada_ruta = "sin meta ruta"
                
                    if metadato_url:
                        # Intenta obtener dc.identifier, dc.identifier.isbn, o dc.identifier.uri
                        for xpath in [
                            "//td[@class='label-cell'][text()='dc.identifier']/following-sibling::td[@class='word-break']",
                            "//td[@class='label-cell'][text()='dc.identifier.isbn']/following-sibling::td[@class='word-break']",
                            "//td[@class='label-cell'][text()='dc.identifier.uri']/following-sibling::td[@class='word-break']"
                        ]:
                            try:
                                dc_identifier = driver.find_element(By.XPATH, xpath).text.replace('/', '|')
                                break
                            except NoSuchElementException:
                                continue
                
                        # Si no se encontró dc.identifier, usa el título como dc.identifier
                        if not dc_identifier:
                            dc_identifier = titulo
                
                        nombre_archivo = f"{normalizar_nombre_archivo(dc_identifier)}.csv"
                        metada_ruta = os.path.join(download_path_mtd, nombre_archivo)
                
                        # Obtener la tabla de metadatos y guardarla en un archivo CSV
                        html = driver.page_source
                        soup = BeautifulSoup(html, 'html.parser')
                        tabla = soup.find('table', class_='ds-includeSet-table')
                        
                        if tabla:
                            with open(metada_ruta, 'w', newline='', encoding='utf-8') as csvfile:
                                csvwriter = csv.writer(csvfile, delimiter='|')
                                for fila in tabla.find_all('tr'):
                                    celdas = fila.find_all('td')
                                    fila_csv = [celda.get_text(strip=True) for celda in celdas]
                                    csvwriter.writerow(fila_csv)
                        else:
                            print("No se encontró la tabla en la página.")
                    else:
                        print("No se encontró el enlace de metadatos en la página.")
                
                except Exception as e:
                    print(f"Se produjo un error: {e}")
                        
                  
                # =============================================================================
                # # Fecha
                # =============================================================================

                    
                try:
                    # Utilizar XPath para encontrar el elemento de la fecha
                    xpath_fecha = "//td[@class='label-cell'][text()='dc.date.issued']/following-sibling::td[@class='word-break']"
                    fecha_contenedor = driver.find_element(By.XPATH, xpath_fecha)
                
                    # Ejecutar el script JavaScript para extraer el texto
                    fecha = driver.execute_script("""
                        var contenedor = arguments[0];
                        var nodos = contenedor.childNodes;
                        var texto = '';
                        for (var i = 0; i < nodos.length; i++) {
                            if (nodos[i].nodeType === Node.TEXT_NODE) {
                                texto += nodos[i].textContent.trim();
                            }
                        }
                        return texto;
                    """, fecha_contenedor)
                except NoSuchElementException:
                    fecha = "No disponible"
                except Exception as e:
                    print(f"Error al obtener la fecha: {e}")
                    fecha = "Error"
                

                 
                # =============================================================================
                # =============================================================================                     
                # =============================================================================
                # # Descargar documentos
                # =============================================================================
                # =============================================================================
                    # Inicializa las listas para los documentos
                ext_docs = []  # Inicializa variables como una lista vacía
                href_doc = []  #  
                doc_urls = []  #  
                status_doc = []  #  
                                
                                    # Resto del código previo...               
                try:
                    # Variable para rastrear si se utilizó el selector alternativo
                    uso_selector_alternativo = False
                
                    elementos_documentos = driver.find_elements(By.CSS_SELECTOR, "div.file-list div.file-link.col-xs-6.col-xs-offset-6.col-sm-2.col-sm-offset-0 a")
                
                    # Si no se encuentran elementos, buscar con el selector alternativo
                    if not elementos_documentos:
                        uso_selector_alternativo = True
                        driver.execute_script("window.history.go(-1)")
                        elementos_documentos = driver.find_elements(By.CSS_SELECTOR, "div.item-page-field-wrapper.table.word-break a")
                
                    for i, elemento in enumerate(elementos_documentos, start=1):
                        doc_url = elemento.get_attribute('href')
                        extension_archivo = obtener_extension(doc_url)
                        doc_nombre_archivo = f"{normalizar_nombre_archivo(dc_identifier)}_doc_{i}"
                        doc_ruta_completa = os.path.join(download_path_doc, doc_nombre_archivo)
                        
                        if not os.path.exists(doc_ruta_completa):
                            # Si el archivo no existe, procede a descargarlo
                            resultado_descarga, estado_descarga = descargar_archivo(doc_url, doc_nombre_archivo, download_path_doc)
                            
                            if resultado_descarga:
                                esperar_descarga_completa(resultado_descarga)
                                href_doc.append(resultado_descarga)
                                ext_docs.append(extension_archivo)
                                doc_urls.append(doc_url)
                                status_doc.append(estado_descarga)
                            else:
                                href_doc.append("no se pudo descargar el archivo")
                                ext_docs.append("")
                                doc_urls.append("")
                                status_doc.append("Error en la descarga")
                        else:
                            href_doc.append(doc_ruta_completa)
                            ext_docs.append(extension_archivo)
                            doc_urls.append(doc_url)
                            status_doc.append("Archivo ya existente")
                    

                
                except (NoSuchElementException, TimeoutException) as e:
                    # Manejo de excepciones
                    href_doc.append("no se pudo procesar el documento")
                    ext_docs.append("")
                    doc_urls.append("")
                    status_doc.append("Error en la búsqueda de documentos")
                
                # =============================================================================
                # Volver a la página principal
                # =============================================================================
                if uso_selector_alternativo:
                    # Retroceder una página si se usó el selector alternativo
                    driver.execute_script("window.history.go(-1)")
                else:
                    # Retroceder dos páginas si se usó el selector original
                    driver.execute_script("window.history.go(-2)")


                # =============================================================================
                # # Agregar la informacion recolectada al libro de Excel
                # =============================================================================


                
                # Asegúrate de que los encabezados se agreguen si el libro está vacío
                if ws.max_row == 1 and ws.max_column == 1:
                    # Agregar los encabezados incluyendo 'Indice' al principio y '%descarga' al final para el estado de la descarga
                    ws.append(['Indice', "dc.identifier", "Titulo", "Autores", "Abstract", "Coleccion", "dc.date.issued", 'URL_Base', 'Url_Repo', "DOI", "Metadato", 'Href_doc', 'Href_doc_2', 'Href_doc_3', 'Href_doc_4','Href_doc_5','Href_doc_6','Href_doc_7','Href_doc_8','Href_doc_9','Href_doc_10','Href_doc_11','Href_doc_12','Href_doc_13','Href_doc_14','Href_doc_15','Href_doc_16','Href_doc_17','Href_doc_18','Href_doc_19','Href_doc_20','Href_doc_21','Href_doc_22','Href_doc_23','Href_doc_24','Href_doc_25','Href_doc_26','Href_doc_27','Href_doc_28','Href_doc_29','Href_doc_30', 'ext_doc', 'ext_doc_2', 'ext_doc_3', 'ext_doc_4','ext_doc_5','ext_doc_6','ext_doc_7','ext_doc_8','ext_doc_9','ext_doc_10','ext_doc_11', 'ext_doc_12', 'ext_doc_13', 'ext_doc_14','ext_doc_15','ext_doc_16','ext_doc_17','ext_doc_18','ext_doc_19','ext_doc_20','ext_doc_21', 'ext_doc_22', 'ext_doc_23', 'ext_doc_24','ext_doc_25','ext_doc_26','ext_doc_27','ext_doc_28','ext_doc_29','ext_doc_30', 'doc_urls', 'doc_urls_2', 'doc_urls_3', 'doc_urls_4','doc_urls_5','doc_urls_6','doc_urls_7','doc_urls_8','doc_urls_9','doc_urls_10','doc_urls_11', 'doc_urls_12', 'doc_urls_13', 'doc_urls_14','doc_urls_15','doc_urls_16','doc_urls_17','doc_urls_18','doc_urls_19','doc_urls_20','doc_urls_21', 'doc_urls_22', 'doc_urls_23', 'doc_urls_24','doc_urls_25','doc_urls_26','doc_urls_27','doc_urls_28','doc_urls_29','doc_urls_30', 'stat_doc', 'stat_doc_2', 'stat_doc_3', 'stat_doc_4','stat_doc_5','stat_doc_6','stat_doc_7','stat_doc_8','stat_doc_9','stat_doc_10','stat_doc_11', 'stat_doc_12', 'stat_doc_13', 'stat_doc_14','stat_doc_15','stat_doc_16','stat_doc_17','stat_doc_18','stat_doc_19','stat_doc_20','stat_doc_21', 'stat_doc_22', 'stat_doc_23', 'stat_doc_24','stat_doc_25','stat_doc_26','stat_doc_27','stat_doc_28','stat_doc_29','stat_doc_30'])
               
                # Verificar si el título ya está en el Excel
                ya_registrado = any(ws.cell(row=i, column=9).value == url_actual for i in range(2, ws.max_row + 1))
                if not ya_registrado:
                    # Agregar la información recolectada al libro de Excel
                    # Asegúrate de que cada lista sea una cadena antes de aplicar split()
                    href_doc_str = ", ".join(href_doc) if isinstance(href_doc, list) else href_doc
                    href_docs = href_doc_str.split(',')
                    
                    ext_docs_str = ", ".join(ext_docs) if isinstance(ext_docs, list) else ext_docs
                    exts = ext_docs_str.split(',')
                    
                    doc_urls_str = ", ".join(doc_urls) if isinstance(doc_urls, list) else doc_urls
                    doc_urls_list = doc_urls_str.split(',')
                    
                    status_doc_str = ", ".join(status_doc) if isinstance(status_doc, list) else status_doc
                    status_list = status_doc_str.split(',')
                    
                    # Agregar la información recolectada al libro de Excel
                    fila_excel = [indice, dc_identifier, titulo, autor, abstract, coleccion, fecha, url_base,url_actual, doi, metada_ruta]
                    
                    # Extiende la fila_excel con la información de los documentos y su estado de descarga
                    fila_excel.extend(href_docs + [''] * (30 - len(href_docs)))
                    fila_excel.extend(exts + [''] * (30 - len(exts)))
                    fila_excel.extend(doc_urls_list + [''] * (30 - len(doc_urls_list)))
                    fila_excel.extend(status_list + [''] * (30 - len(status_list)))
                    
                    ws.append(fila_excel)
                    
                    indice += 1
                
                
# =============================================================================
#                 # Ajuste
# =============================================================================
                for col_num in range(10, 10 + len(href_docs) + len(exts) + len(doc_urls_list)):
                    ws.column_dimensions[get_column_letter(col_num)].width = 20


# =============================================================================
#                iteraciones
# =============================================================================

                contador_iteraciones += 1
                if contador_iteraciones % 20 == 0:
                    print("Pausa de 1 minuto para evitar sobrecarga del servidor.")
                    time.sleep(60) 
                    
# =============================================================================
#               Save excel   
# ============================================================================                
           
            wb.save("Planilla_Proyecto_InES.xlsx")  
            wb.close()
            if offset >= 2999:
                break  # Salir del bucle si hemos alcanzado el li­mite
            
            
            # Incrementa el offset en 20
            offset += 20 
            

    except (NoSuchElementException, TimeoutException,TimeoutError) as e:
        
        print(f"Error al intentar encontrar un elemento: {e}")
        raise ServerError("Error relacionado con el servidor o la página web.")
        
    except UnboundLocalError as e:
        print(f"Error de variable local no asignada: {e}")
        raise ServerError("Error de variable local no asignada.")   

    finally:
        if driver is not None:
            driver.quit()
        if wb:
            wb.save("Planilla_Proyecto_InES.xlsx")  # Guarda los cambios finales
            wb.close()
            

# =============================================================================
# =============================================================================
# # Bucle principal para manejar el reinicio del navegador
# =============================================================================
# =============================================================================
            
archivo_excel = 'Planilla_Proyecto.xlsx'
offset_inicial = 0
if os.path.exists(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active
    if ws.max_row > 1 and ws.cell(row=ws.max_row, column=8).hyperlink:
        offset_inicial = obtener_offset_del_enlace(ws.cell(row=ws.max_row, column=8).hyperlink.target)



max_retries = 823  # Total de intentos durante 4 días
retry_count = 0
wait_time = 420  # Tiempo de espera en segundos (7 minutos)

while retry_count < max_retries:
    try:
        login_and_scrape(offset_inicial)
        break  # Salir del bucle si se ejecuta con éxito
    except ServerError:
        print(f"Problemas con el servidor o la página web. Reintentando... ({retry_count + 1}/{max_retries})")
        retry_count += 1
        if retry_count < max_retries:
            print("Esperando 7 minutos para reintentar...")
            time.sleep(wait_time)  # Espera de 7 minutos antes de reintentar
    except NoSuchWindowException:
        print("El navegador se ha cerrado inesperadamente. Finalizando el proceso.")
 